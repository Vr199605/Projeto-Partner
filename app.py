# app.py
import re
import io
import base64
import unicodedata
from datetime import datetime

import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.colors import HexColor
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, PageBreak)
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart

st.set_page_config(page_title="Dashboard Financeiro Premium", layout="wide", page_icon="📊")

# ---------------------------------------------------------------------------
# Paleta — UI (cartões/cabeçalhos, sempre com fundo próprio) vs. Gráficos
# (precisam ficar visíveis tanto no tema claro quanto escuro do Streamlit)
# ---------------------------------------------------------------------------
C_NAVY = "#0a1628"
C_NAVY2 = "#1a3a5c"
C_TEAL = "#2e86ab"
C_TEXT_DARK = "#0a1628"
C_GOLD_MEDAL, C_SILVER, C_BRONZE = "#ffd700", "#c0c0c0", "#cd7f32"

# cores "de dado" — vivas o bastante para não sumir em nenhum dos dois temas
CH_BLUE = "#4fb3e8"
CH_BLUE_LIGHT = "#7dd3fc"
CH_GOLD = "#feca57"
CH_SUCCESS = "#00d4aa"
CH_DANGER = "#ff6b6b"
CH_DANGER_LIGHT = "#ffb4b4"
CH_PURPLE = "#a78bfa"
CH_PINK = "#f472b6"
CH_QUALITATIVE = [CH_BLUE, CH_GOLD, CH_SUCCESS, CH_PURPLE, CH_PINK, CH_DANGER, CH_BLUE_LIGHT]

MESES_ORDEM = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
MES_PREFIX = {"JAN": "Jan", "FEV": "Fev", "MAR": "Mar", "ABR": "Abr", "MAI": "Mai", "JUN": "Jun",
              "JUL": "Jul", "AGO": "Ago", "SET": "Set", "OUT": "Out", "NOV": "Nov", "DEZ": "Dez"}
QUARTER_MONTHS = {"1º Tri": ["Jan", "Fev", "Mar"], "2º Tri": ["Abr", "Mai", "Jun"],
                   "3º Tri": ["Jul", "Ago", "Set"], "4º Tri": ["Out", "Nov", "Dez"]}
QUARTER_OF = {m: q for q, ms in QUARTER_MONTHS.items() for m in ms}

# ---------------------------------------------------------------------------
# CSS — todo elemento com fundo próprio define sua cor de texto explicitamente,
# para não herdar a cor de texto do tema (evita texto invisível no modo escuro)
# ---------------------------------------------------------------------------
st.markdown(f"""
<style>
@keyframes gradientShift {{
    0% {{ background-position: 0% 50%; }}
    50% {{ background-position: 100% 50%; }}
    100% {{ background-position: 0% 50%; }}
}}
.hero {{ background: linear-gradient(135deg, {C_NAVY} 0%, {C_NAVY2} 40%, {C_TEAL} 100%);
    background-size: 200% 200%; animation: gradientShift 8s ease infinite;
    padding: 28px 32px; border-radius: 16px; color: white; margin-bottom: 20px;
    box-shadow: 0 8px 24px rgba(10,22,40,.35); }}
.hero h1 {{ margin:0; font-size: 27px; color:white; }}
.hero p {{ margin:4px 0 0 0; opacity:.9; font-size:13px; color:white; }}
.kpi-card {{ border-radius: 16px; padding: 18px 16px; text-align:center; background:#ffffff;
    color:{C_TEXT_DARK}; box-shadow: 0 4px 16px rgba(10,22,40,.10); border:1px solid #eef1f5;
    transition:.15s; }}
.kpi-card:hover {{ box-shadow: 0 10px 24px rgba(10,22,40,.16); transform: translateY(-2px); }}
.kpi-label {{ font-size:11px; font-weight:800; color:#5a6b7d; letter-spacing:.6px; text-transform:uppercase;}}
.kpi-value {{ font-size:24px; font-weight:800; color:{C_TEXT_DARK}; margin:6px 0 2px 0;}}
.kpi-delta-up {{ font-size:12px; font-weight:700; color:#00916e; }}
.kpi-delta-down {{ font-size:12px; font-weight:700; color:#d9364a; }}
.section-title {{ font-size:19px; font-weight:800; color: var(--text-color, {C_NAVY2});
    margin:22px 0 12px 0; border-bottom:3px solid {C_TEAL}; padding-bottom:6px; }}
.legenda-box {{ background:#f4f8fb; border-left:4px solid {C_TEAL}; border-radius:8px;
    padding:12px 16px; font-size:12.5px; color:#33475b; margin-bottom:6px; }}
.filtro-box {{ background:linear-gradient(135deg,{C_NAVY} 0%,{C_NAVY2} 100%); border-radius:14px;
    padding:16px 20px; color:white; margin-bottom:14px; }}
.filtro-box b, .filtro-box span {{ color:white; }}
.socio-card {{ border-radius:16px; padding:18px; color:white; box-shadow:0 4px 16px rgba(10,22,40,.15);}}
.socio-card b, .socio-card span, .socio-card h2 {{ color:white; }}
.rank-card {{ border-radius:14px; padding:14px 16px; display:flex; align-items:center; gap:14px;
    box-shadow:0 3px 12px rgba(10,22,40,.08); margin-bottom:10px; background:white; color:{C_TEXT_DARK}; }}
.rank-medal {{ font-size:30px; }}
.total-box {{ border-radius:12px; padding:14px 18px; font-weight:700; font-size:15px; margin-top:8px; }}
.empty-state {{ background:linear-gradient(135deg, rgba(46,134,171,.07), rgba(10,22,40,.03));
    border:1.5px dashed {C_TEAL}; border-radius:14px; padding:26px; text-align:center;
    color:#5a6b7d; margin-bottom:8px; }}
.empty-state-icon {{ font-size:30px; margin-bottom:8px; }}
.empty-state-msg {{ font-size:13px; }}

/* ---------------- Storytelling tab ---------------- */
.story-hero {{ background: linear-gradient(120deg, {C_NAVY} 0%, {C_NAVY2} 45%, {C_TEAL} 100%);
    background-size: 200% 200%; animation: gradientShift 10s ease infinite;
    border-radius:18px; padding:26px 30px; color:white; margin-bottom:18px;
    box-shadow: 0 10px 28px rgba(10,22,40,.30); }}
.story-hero .story-kicker {{ font-size:11.5px; font-weight:800; letter-spacing:1.2px;
    text-transform:uppercase; opacity:.85; color:#dff0ff; }}
.story-hero h2 {{ margin:6px 0 8px 0; font-size:24px; color:white; line-height:1.35; }}
.story-hero p {{ margin:0; font-size:13.5px; color:#eaf3fb; opacity:.95; line-height:1.5; }}
.story-section-title {{ font-size:17px; font-weight:800; color: var(--text-color, {C_NAVY2});
    margin:26px 0 10px 0; display:flex; align-items:center; gap:8px; }}
.story-p {{ font-size:13.5px; line-height:1.75; color:#33475b; margin:0 0 10px 0; }}
.story-grid {{ display:grid; grid-template-columns:repeat(auto-fit, minmax(220px,1fr)); gap:12px; margin-bottom:14px; }}
.story-card {{ border-radius:14px; padding:16px 18px; background:#ffffff; color:{C_TEXT_DARK};
    box-shadow:0 4px 14px rgba(10,22,40,.08); border-left:5px solid {C_TEAL}; }}
.story-card.good {{ border-left-color:{CH_SUCCESS}; }}
.story-card.warn {{ border-left-color:{CH_GOLD}; }}
.story-card.bad {{ border-left-color:{CH_DANGER}; }}
.story-card-title {{ font-size:12px; font-weight:800; text-transform:uppercase; letter-spacing:.4px;
    color:#5a6b7d; margin-bottom:6px; }}
.story-card-value {{ font-size:20px; font-weight:800; margin-bottom:4px; }}
.story-card-note {{ font-size:12px; color:#6b7a8d; line-height:1.5; }}
.callout {{ border-radius:12px; padding:14px 18px; font-size:13px; line-height:1.6; margin:10px 0 16px 0;
    border:1px solid rgba(0,0,0,.06); }}
.callout.info {{ background:rgba(46,134,171,.08); border-left:4px solid {C_TEAL}; color:#1c3a52; }}
.callout.good {{ background:rgba(0,212,170,.10); border-left:4px solid {CH_SUCCESS}; color:#0a5c47; }}
.callout.warn {{ background:rgba(254,202,87,.14); border-left:4px solid {CH_GOLD}; color:#6b4e00; }}
.callout.bad {{ background:rgba(255,107,107,.10); border-left:4px solid {CH_DANGER}; color:#7a1f24; }}
.callout b {{ color: inherit; }}
.reco-item {{ display:flex; gap:12px; align-items:flex-start; padding:10px 4px; border-bottom:1px dashed #e3e8ee; }}
.reco-item:last-child {{ border-bottom:none; }}
.reco-num {{ flex:0 0 auto; width:26px; height:26px; border-radius:50%; background:{C_TEAL}; color:white;
    font-weight:800; font-size:12.5px; display:flex; align-items:center; justify-content:center; }}
.reco-text {{ font-size:13px; line-height:1.6; color:#33475b; }}
.reco-text b {{ color:{C_NAVY2}; }}

/* ---------------- Apresentação Executiva ---------------- */
.exec-hero {{ background: linear-gradient(120deg, {C_NAVY} 0%, {C_NAVY2} 55%, {C_TEAL} 100%);
    background-size: 200% 200%; animation: gradientShift 12s ease infinite;
    border-radius:20px; padding:30px 34px; color:white; margin-bottom:22px;
    box-shadow: 0 12px 32px rgba(10,22,40,.32); }}
.exec-hero .exec-kicker {{ font-size:12px; font-weight:800; letter-spacing:1.6px; text-transform:uppercase;
    opacity:.85; color:#dff0ff; }}
.exec-hero h1 {{ margin:8px 0 0 0; font-size:28px; color:white; }}
.exec-hero p {{ margin:6px 0 0 0; font-size:13.5px; color:#eaf3fb; opacity:.9; }}
.exec-tile {{ border-radius:16px; padding:20px 18px; text-align:center; background:#ffffff;
    color:{C_TEXT_DARK}; box-shadow:0 6px 20px rgba(10,22,40,.10); border:1px solid #eef1f5; }}
.exec-tile-label {{ font-size:11.5px; font-weight:800; color:#5a6b7d; letter-spacing:.6px; text-transform:uppercase; }}
.exec-tile-value {{ font-size:27px; font-weight:800; color:{C_TEXT_DARK}; margin:8px 0 2px 0; }}
.exec-slide {{ background:#ffffff; border-radius:18px; padding:24px 26px; margin-bottom:22px;
    box-shadow:0 4px 18px rgba(10,22,40,.07); border:1px solid #f0f2f5; }}
.exec-slide-num {{ display:inline-flex; align-items:center; justify-content:center; width:28px; height:28px;
    border-radius:50%; background:{C_TEAL}; color:white; font-weight:800; font-size:13px; margin-right:10px; }}
.exec-slide-title {{ font-size:18px; font-weight:800; color:{C_NAVY2}; display:flex; align-items:center;
    margin-bottom:4px; }}
.exec-headline {{ font-size:14.5px; font-weight:600; color:#33475b; line-height:1.55; margin:6px 0 16px 44px; }}
.exec-headline b {{ color:{C_NAVY2}; }}
</style>
""", unsafe_allow_html=True)

PLOTLY_CONFIG = {"displaylogo": False, "modeBarButtonsToRemove": ["lasso2d", "select2d"],
                  "scrollZoom": False,
                  "toImageButtonOptions": {"format": "png", "scale": 2}}

def empty_state(icon, message):
    st.markdown(f"""<div class="empty-state"><div class="empty-state-icon">{icon}</div>
    <div class="empty-state-msg">{message}</div></div>""", unsafe_allow_html=True)

def style_fig(fig, height=None, hovermode="closest"):
    """Deixa o gráfico legível tanto no tema claro quanto no escuro do Streamlit:
    fundo transparente + grade/texto em cinza neutro (bom contraste nos dois temas).
    Também aplica camada de interatividade: hover unificado com tooltip estilizado,
    spikes (crosshair) nos eixos e transições suaves de entrada/atualização."""
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#8a99ab", size=12, family="'Segoe UI', Inter, -apple-system, sans-serif"),
        legend=dict(font=dict(color="#8a99ab"), bgcolor="rgba(0,0,0,0)",
                    orientation="h", y=-0.18, x=0.5, xanchor="center"),
        margin=dict(t=50, b=40, l=10, r=10),
        hovermode=hovermode,
        hoverlabel=dict(bgcolor=C_NAVY2, font_color="white", font_size=12.5,
                         bordercolor=C_TEAL, font_family="'Segoe UI', sans-serif"),
        hoverdistance=40,
        transition=dict(duration=350, easing="cubic-in-out"),
        uniformtext=dict(minsize=9, mode="hide"),
    )
    fig.update_xaxes(gridcolor="rgba(140,150,160,.22)", zerolinecolor="rgba(140,150,160,.45)",
                      linecolor="rgba(140,150,160,.4)", showspikes=True, spikemode="across",
                      spikesnap="cursor", spikethickness=1, spikecolor=CH_BLUE, spikedash="dot")
    fig.update_yaxes(gridcolor="rgba(140,150,160,.22)", zerolinecolor="rgba(140,150,160,.45)",
                      linecolor="rgba(140,150,160,.4)")
    if height:
        fig.update_layout(height=height)
    return fig

def finish_hbar(fig, values, height=None):
    """Para rankings horizontais com rótulo de valor 'outside': abre espaço extra no
    eixo X e na margem direita para o texto não ser cortado pelo container."""
    fig = style_fig(fig, height=height)
    vmax = max(values) if len(values) else 0
    fig.update_xaxes(range=[0, vmax * 1.24 if vmax else 1])
    fig.update_layout(margin=dict(r=80))
    return fig

# ---------------------------------------------------------------------------
# Helpers gerais
# ---------------------------------------------------------------------------
def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFKD", str(s)) if not unicodedata.combining(c))

def norm_header(s):
    return strip_accents(str(s)).upper().strip()

def parse_number(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return 0.0 if pd.isna(v) else float(v)
    s = str(v).strip()
    if s in ("", "-", "nan", "None"):
        return 0.0
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = s.replace("R$", "").replace(" ", "")
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        val = float(s)
    except ValueError:
        return 0.0
    return -val if neg else val

def find_col(df, *keywords, exclude=None):
    for col in df.columns:
        h = norm_header(col)
        if all(k in h for k in keywords) and (not exclude or exclude not in h):
            return col
    return None

def find_sheet(sheetnames, *contains):
    for name in sheetnames:
        h = norm_header(name)
        if any(c in h for c in contains):
            return name
    return None

def month_from_value(v):
    """Converte um valor de data/mês em 'Jan'..'Dez'. Trata None/NaN/NaT com segurança."""
    try:
        if v is None or pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, (pd.Timestamp, datetime)):
        return MESES_ORDEM[v.month - 1]
    s = str(v).strip()
    m = re.match(r"^(\d{1,2})[/-](\d{4})$", s)
    if m:
        return MESES_ORDEM[int(m.group(1)) - 1]
    try:
        dt = pd.to_datetime(v, dayfirst=True, errors="raise")
        if pd.isna(dt):
            return None
        return MESES_ORDEM[dt.month - 1]
    except Exception:
        return None

def detect_month_name(sheet_name):
    clean = re.sub(r"[^A-Za-zÀ-ÿ]", "", sheet_name).upper()[:3].replace("Ç", "C")
    return MES_PREFIX.get(clean)

# ---------------------------------------------------------------------------
# Parsing "DRE 2026"
# ---------------------------------------------------------------------------
def _norm(v):
    return str(v).strip().upper() if v is not None else ""

def find_row(ws, label, start=1, end=None):
    end = end or ws.max_row
    target = _norm(label)
    for r in range(start, end + 1):
        if target in _norm(ws.cell(row=r, column=2).value):
            return r
    return None

def find_header_row(ws):
    for r in range(1, ws.max_row + 1):
        vals = [_norm(ws.cell(row=r, column=c).value)[:3] for c in range(3, ws.max_column + 1)]
        if sum(1 for v in vals if v in [m.upper() for m in MES_PREFIX.values()]) >= 6:
            return r
    return None

def row_series(ws, row, cols):
    return [ws.cell(row=row, column=c).value or 0 for c in cols]

def parse_dre(ws):
    header_row = find_header_row(ws)
    if header_row is None:
        return None
    month_cols = []
    for c in range(3, ws.max_column + 1):
        v = _norm(ws.cell(row=header_row, column=c).value)
        if v[:3] in [m.upper() for m in MES_PREFIX.values()]:
            month_cols.append(c)
    if not month_cols:
        return None

    direta_sec = find_row(ws, "PRODUÇÃO DIRETA")
    portal_sec = find_row(ws, "PORTAL MAAS")
    resultado_row = find_row(ws, "RESULTADO OPERACIONAL", start=portal_sec)
    while resultado_row and "DISTRIBUI" in _norm(ws.cell(row=resultado_row, column=2).value):
        resultado_row = find_row(ws, "RESULTADO OPERACIONAL", start=resultado_row + 1)

    d_end, p_end = portal_sec, resultado_row
    rows = {
        "d_receita": find_row(ws, "RECEITA BRUTA", direta_sec, d_end),
        "d_impostos": find_row(ws, "IMPOSTOS DIRETOS", direta_sec, d_end),
        "d_custo": find_row(ws, "CUSTO OPERACIONAL", direta_sec, d_end),
        "d_cocorretagem": find_row(ws, "CO-CORRETAGEM", direta_sec, d_end),
        "d_rebate": find_row(ws, "REBATE AAI", direta_sec, d_end),
        "d_margem": find_row(ws, "MARGEM DE CONTRIBUIÇÃO", direta_sec, d_end),
        "d_despesas": find_row(ws, "DESPESAS", direta_sec, d_end),
        "d_folha": find_row(ws, "FOLHA+TERCEIROS", direta_sec, d_end),
        "d_ebitda": find_row(ws, "EBITDA SOCIETÁRIO", direta_sec, d_end),
        "p_receita": find_row(ws, "RECEITA BRUTA", portal_sec, p_end),
        "p_impostos": find_row(ws, "IMPOSTOS DIRETOS", portal_sec, p_end),
        "p_custo": find_row(ws, "CUSTO OPERACIONAL", portal_sec, p_end),
        "p_margem": find_row(ws, "MARGEM DE CONTRIBUIÇÃO", portal_sec, p_end),
        "p_despesas": find_row(ws, "DESPESAS", portal_sec, p_end),
        "p_folha": find_row(ws, "FOLHA+TERCEIROS", portal_sec, p_end),
        "p_ebitda": find_row(ws, "EBITDA SOCIETÁRIO", portal_sec, p_end),
    }
    socio_partner_row = find_row(ws, "Sócio Partner", start=resultado_row)
    socio_maldivas_row = find_row(ws, "Sócio Maldivas", start=resultado_row)
    valor_pagar_maldivas_row = find_row(ws, "Valor a pagar", start=resultado_row)

    def s(key):
        r = rows.get(key)
        return row_series(ws, r, month_cols) if r else [0] * len(month_cols)

    def rs(row):
        return row_series(ws, row, month_cols) if row else [0] * len(month_cols)

    meses = [MES_PREFIX[_norm(ws.cell(row=header_row, column=c).value)[:3]] for c in month_cols]

    receita_direta = s("d_receita")
    receita_portal = s("p_receita")
    df = pd.DataFrame({
        "ReceitaDireta": receita_direta, "ReceitaPortal": receita_portal,
        "Impostos": [a + b for a, b in zip(s("d_impostos"), s("p_impostos"))],
        "Custo": [a + b for a, b in zip(s("d_custo"), s("p_custo"))],
        "CoCorretagem": s("d_cocorretagem"), "RebateAAI": s("d_rebate"),
        "MargemContribuicao": [a + b for a, b in zip(s("d_margem"), s("p_margem"))],
        "Despesas": [a + b for a, b in zip(s("d_despesas"), s("p_despesas"))],
        "Folha": [a + b for a, b in zip(s("d_folha"), s("p_folha"))],
        "EBITDA_Direta": s("d_ebitda"), "EBITDA_Portal": s("p_ebitda"),
        "ResultadoOperacional": rs(resultado_row),
        "SocioPartner": rs(socio_partner_row), "SocioMaldivas": rs(socio_maldivas_row),
        "ValorPagarMaldivas": rs(valor_pagar_maldivas_row),
    }, index=meses)
    df["TemDados"] = ((df["ReceitaDireta"] + df["ReceitaPortal"]) > 0).astype(int)
    df = df.groupby(df.index).sum()
    return df.reindex([m for m in MESES_ORDEM if m in df.index])

def parse_shares(wb):
    if "INPUTS" not in wb.sheetnames:
        return 0.7, 0.3
    ws = wb["INPUTS"]
    rp, rm = find_row(ws, "Sócio Partner"), find_row(ws, "Sócio Maldivas")
    partner = ws.cell(row=rp, column=2).value if rp else 0.7
    maldivas = ws.cell(row=rm, column=2).value if rm else 0.3
    return float(partner or 0.7), float(maldivas or 0.3)

# ---------------------------------------------------------------------------
# Parsing de transações — caminho padrão: aba única (ex. "ASSERTIF DIRETO")
# ---------------------------------------------------------------------------
def parse_transacoes(file_bytes, sheet_name):
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
    except Exception:
        return None
    if df is None or df.empty:
        return None

    col_data = find_col(df, "DATA") or find_col(df, "MES")
    col_cliente = find_col(df, "CLIENTE")
    col_seguradora = find_col(df, "SEGURADORA")
    col_produto = find_col(df, "PRODUTO")
    col_originador = find_col(df, "ORIGINADOR")
    col_comissao = find_col(df, "COMISS") or find_col(df, "VALOR")
    if col_comissao is None:
        return None

    out = pd.DataFrame()
    out["Mes"] = df[col_data].apply(month_from_value) if col_data is not None else None
    out["Cliente"] = df[col_cliente].fillna("Não informado").astype(str) if col_cliente is not None else "Não informado"
    out["Seguradora"] = df[col_seguradora].fillna("Não informado").astype(str) if col_seguradora is not None else "Não informado"
    out["Produto"] = df[col_produto].fillna("Não informado").astype(str) if col_produto is not None else "Não informado"
    out["Originador"] = df[col_originador].fillna("Não informado").astype(str) if col_originador is not None else "Não informado"
    out["Valor"] = df[col_comissao].apply(parse_number)
    out = out[out["Valor"] != 0]
    return out if not out.empty else None

def parse_transacoes_generic(file_bytes, sheet_name):
    """Como parse_transacoes, mas só aceita a aba se ela tiver ao menos uma coluna
    de dimensão (Cliente/Seguradora/Produto/Originador) — usado apenas na varredura
    genérica de abas não reconhecidas, para não confundir uma aba de despesas (que
    também tem uma coluna de valor) com uma aba de receita."""
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
    except Exception:
        return None
    if df is None or df.empty:
        return None
    tem_dimensao = any(find_col(df, k) is not None for k in ("CLIENTE", "SEGURADORA", "PRODUTO", "ORIGINADOR"))
    if not tem_dimensao:
        return None
    return parse_transacoes(file_bytes, sheet_name)

# ---------------------------------------------------------------------------
# EXCEÇÃO — RJ+2026: sem aba consolidada; Seguradora/Produto/Originador/Cliente
# vêm das abas mensais (JAN_2026, FEV_2026...). Mapeamento assumido:
#   coluna "PRODUTO"   (ex.: AMIL)               -> nossa "Seguradora"
#   coluna "CATEGORIA" (ex.: Seguros Corporativos) -> nosso "Produto"
#   "CÓD. ASSESSOR"    + aba BASE (COD INTERNO -> NOME ASSESSOR) -> "Originador"
#   "NOME CLIENTE"                                -> "Cliente"
#   "COMISSÃO BRUTA D.A" (exclui "PARTNER"/"RECEBIDA")           -> "Valor"
# ---------------------------------------------------------------------------
def parse_transacoes_from_monthly(wb):
    code_to_name = {}
    if "BASE" in wb.sheetnames:
        base_ws = wb["BASE"]
        code_col = name_col = None
        for c in range(1, base_ws.max_column + 1):
            h = _norm(base_ws.cell(row=1, column=c).value)
            if "COD INTERNO" in h:
                code_col = c
            if "NOME ASSESSOR" in h:
                name_col = c
        if code_col and name_col:
            for r in range(2, base_ws.max_row + 1):
                code = base_ws.cell(row=r, column=code_col).value
                name = base_ws.cell(row=r, column=name_col).value
                if code and name and str(code).strip() not in code_to_name:
                    code_to_name[str(code).strip()] = str(name).strip()

    records = []
    for sheet_name in wb.sheetnames:
        mes = detect_month_name(sheet_name)
        if not mes:
            continue
        ws = wb[sheet_name]
        col_assessor = col_cliente = col_categoria = col_produto = col_valor = None
        for c in range(1, ws.max_column + 1):
            h = _norm(ws.cell(row=2, column=c).value)
            if "ASSESSOR" in h and "COD" in h:
                col_assessor = c
            if "NOME CLIENTE" in h:
                col_cliente = c
            if h == "CATEGORIA":
                col_categoria = c
            if h == "PRODUTO":
                col_produto = c
            if "COMISS" in h and "D.A" in h and "PARTNER" not in h and "RECEBIDA" not in h:
                col_valor = c
        if not col_valor:
            continue
        for r in range(3, ws.max_row + 1):
            val = ws.cell(row=r, column=col_valor).value
            if not val:
                continue
            code = ws.cell(row=r, column=col_assessor).value if col_assessor else None
            cliente = ws.cell(row=r, column=col_cliente).value if col_cliente else None
            categoria = ws.cell(row=r, column=col_categoria).value if col_categoria else None
            produto = ws.cell(row=r, column=col_produto).value if col_produto else None
            originador = code_to_name.get(str(code).strip(), str(code).strip()) if code else "Não informado"
            records.append((mes, str(cliente).strip() if cliente else "Não informado",
                             str(produto).strip() if produto else "Não informado",
                             str(categoria).strip() if categoria else "Não informado",
                             originador, float(val)))
    if not records:
        return None
    return pd.DataFrame(records, columns=["Mes", "Cliente", "Seguradora", "Produto", "Originador", "Valor"])

def parse_despesas(file_bytes, sheet_name):
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
    except Exception:
        return None
    if df is None or df.empty:
        return None
    col_data = find_col(df, "DATA") or find_col(df, "MES")
    col_categoria = find_col(df, "CATEGORIA")
    col_valor = find_col(df, "VALOR")
    if col_valor is None or col_categoria is None:
        return None
    out = pd.DataFrame()
    out["Mes"] = df[col_data].apply(month_from_value) if col_data is not None else None
    out["Categoria"] = df[col_categoria].fillna("Não informado").astype(str)
    out["Valor"] = df[col_valor].apply(parse_number).abs()
    out = out[out["Valor"] != 0]
    return out if not out.empty else None

@st.cache_data(show_spinner="Lendo todas as abas da planilha...")
def parse_workbook(file_bytes, file_name):
    """Lê o workbook inteiro: exige uma aba de DRE (backbone dos KPIs), tenta os
    caminhos conhecidos de receita/despesa e, por fim, varre TODAS as abas que
    sobraram em busca de dados reconhecíveis — para nenhuma informação lançada
    em uma aba com nome fora do padrão ficar de fora do dashboard."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet_dre = find_sheet(wb.sheetnames, "DRE")
    if sheet_dre is None:
        return None
    df = parse_dre(wb[sheet_dre])
    if df is None:
        return None
    shares = parse_shares(wb)

    usadas = {sheet_dre}
    if "INPUTS" in wb.sheetnames:
        usadas.add("INPUTS")
    if "BASE" in wb.sheetnames:
        usadas.add("BASE")

    tx_fontes, desp_fontes = [], []

    sheet_receitas = find_sheet(wb.sheetnames, "DIRETO", "RECEITA")
    tx = parse_transacoes(file_bytes, sheet_receitas) if sheet_receitas else None
    if tx is not None:
        tx_fontes.append(sheet_receitas)
        usadas.add(sheet_receitas)

    if tx is None:
        meses_mensais = [s for s in wb.sheetnames if detect_month_name(s)]
        tx = parse_transacoes_from_monthly(wb)  # exceção RJ+2026: abas mensais (JAN_2026, FEV_2026...)
        if tx is not None:
            tx_fontes.extend(meses_mensais)
            usadas.update(meses_mensais)

    sheet_despesas = find_sheet(wb.sheetnames, "DESPES")
    desp = parse_despesas(file_bytes, sheet_despesas) if sheet_despesas else None
    if desp is not None:
        desp_fontes.append(sheet_despesas)
        usadas.add(sheet_despesas)

    # Varredura genérica: qualquer aba restante que pareça ter dados de despesa
    # ou de receita/produção entra automaticamente, mesmo com nome fora do padrão.
    extra_tx, extra_desp, ignoradas = [], [], []
    for nome in wb.sheetnames:
        if nome in usadas:
            continue
        d = parse_despesas(file_bytes, nome)
        if d is not None:
            extra_desp.append(d)
            desp_fontes.append(nome)
            usadas.add(nome)
            continue
        t = parse_transacoes_generic(file_bytes, nome)
        if t is not None:
            extra_tx.append(t)
            tx_fontes.append(nome)
            usadas.add(nome)
            continue
        ignoradas.append(nome)

    if extra_tx:
        tx = pd.concat([tx] + extra_tx, ignore_index=True) if tx is not None else pd.concat(extra_tx, ignore_index=True)
    if extra_desp:
        desp = pd.concat([desp] + extra_desp, ignore_index=True) if desp is not None else pd.concat(extra_desp, ignore_index=True)

    sheets_info = {
        "dre": sheet_dre, "transacoes": tx_fontes, "despesas": desp_fontes,
        "ignoradas": ignoradas, "total_abas": len(wb.sheetnames),
    }
    return {"file_name": file_name, "df": df, "shares": shares, "tx": tx, "desp": desp, "sheets_info": sheets_info}

def months_with_data(parsed_list, selected_files):
    s = set()
    for p in parsed_list:
        if p["file_name"] in selected_files:
            for m in p["df"].index:
                if p["df"].loc[m, "TemDados"] > 0:
                    s.add(m)
    return sorted(s, key=lambda m: MESES_ORDEM.index(m))

def aggregate(parsed_list, selected_files, selected_months):
    sel = [p for p in parsed_list if p["file_name"] in selected_files]
    if not sel:
        return None
    parts = [p["df"].loc[p["df"].index.intersection(selected_months)] for p in sel]
    combined_full = pd.concat(parts).groupby(level=0).sum()
    combined_full["TemDados"] = combined_full["TemDados"] > 0
    combined_full = combined_full.reindex([m for m in MESES_ORDEM if m in combined_full.index])
    combined = combined_full[combined_full["TemDados"]].copy()

    tx_parts = [p["tx"] for p in sel if p["tx"] is not None]
    combined_tx = pd.concat(tx_parts, ignore_index=True) if tx_parts else None
    if combined_tx is not None:
        combined_tx = combined_tx[combined_tx["Mes"].isna() | combined_tx["Mes"].isin(selected_months)]
        if combined_tx.empty:
            combined_tx = None

    desp_parts = [p["desp"] for p in sel if p["desp"] is not None]
    combined_desp = pd.concat(desp_parts, ignore_index=True) if desp_parts else None
    if combined_desp is not None:
        combined_desp = combined_desp[combined_desp["Mes"].isna() | combined_desp["Mes"].isin(selected_months)]
        if combined_desp.empty:
            combined_desp = None

    partner_share = float(np.mean([p["shares"][0] for p in sel]))
    maldivas_share = float(np.mean([p["shares"][1] for p in sel]))
    return combined, combined_tx, combined_desp, (partner_share, maldivas_share)

def month_delta(series):
    if len(series) < 2:
        return None
    prev, curr = series.iloc[-2], series.iloc[-1]
    if prev == 0:
        return None
    return (curr - prev) / abs(prev)

def delta_html(delta):
    if delta is None:
        return '<span style="font-size:12px;color:#aaa;">sem comparação</span>'
    cls = "kpi-delta-up" if delta >= 0 else "kpi-delta-down"
    arrow = "▲" if delta >= 0 else "▼"
    return f'<span class="{cls}">{arrow} {delta:+.1%} vs mês anterior</span>'

# ---------------------------------------------------------------------------
# Upload
# ---------------------------------------------------------------------------
st.sidebar.header("📁 Upload")
uploads = st.sidebar.file_uploader("Envie uma ou mais planilhas (.xlsx)", type=["xlsx"], accept_multiple_files=True)
logo_file = st.sidebar.file_uploader("Logo da empresa (opcional)", type=["png", "jpg", "jpeg"])
if logo_file is not None:
    st.session_state["logo_bytes"] = logo_file.getvalue()

if not uploads:
    st.info("Envie uma ou mais planilhas contendo uma aba de DRE (ex.: 'DRE 2026') para gerar o dashboard.")
    st.stop()

parsed_list = []
for f in uploads:
    data = parse_workbook(f.getvalue(), f.name)
    if data is None:
        st.sidebar.warning(f"'{f.name}' não contém uma aba de DRE válida — ignorado.")
        continue
    parsed_list.append(data)

if not parsed_list:
    st.error("Nenhum arquivo válido foi encontrado.")
    st.stop()

all_files = [p["file_name"] for p in parsed_list]

with st.sidebar.expander("📄 Abas lidas por arquivo", expanded=False):
    for p in parsed_list:
        info = p["sheets_info"]
        st.markdown(f"**{p['file_name']}** — {info['total_abas']} aba(s)")
        st.caption(f"DRE: `{info['dre']}`")
        st.caption("Receita/produção: " + (", ".join(f"`{s}`" for s in info["transacoes"]) if info["transacoes"] else "nenhuma aba reconhecida"))
        st.caption("Despesas: " + (", ".join(f"`{s}`" for s in info["despesas"]) if info["despesas"] else "nenhuma aba reconhecida"))
        if info["ignoradas"]:
            st.caption("Ignoradas (sem colunas reconhecíveis): " + ", ".join(f"`{s}`" for s in info["ignoradas"]))

st.sidebar.header("🔎 Planilhas")
if "sel_files" not in st.session_state:
    st.session_state.sel_files = all_files
ffc1, ffc2 = st.sidebar.columns(2)
if ffc1.button("Todas as planilhas"):
    st.session_state.sel_files = all_files
if ffc2.button("Limpar planilhas"):
    st.session_state.sel_files = []
selected_files = st.sidebar.multiselect("Planilhas incluídas", all_files, key="sel_files")
mil_mode = st.sidebar.checkbox("Exibir valores em R$ mil")

if not selected_files:
    st.warning("Selecione ao menos uma planilha.")
    st.stop()

available_months = months_with_data(parsed_list, selected_files)
if not available_months:
    st.warning("Nenhum mês com dados lançados foi encontrado nas planilhas selecionadas.")
    st.stop()
quarters_available = [q for q in ["1º Tri", "2º Tri", "3º Tri", "4º Tri"]
                       if any(m in available_months for m in QUARTER_MONTHS[q])]

if "sel_months" not in st.session_state or not set(st.session_state.sel_months).issubset(set(available_months)):
    st.session_state.sel_months = available_months
if "sel_quarters" not in st.session_state:
    st.session_state.sel_quarters = quarters_available

def sync_quarter_to_months():
    qs = st.session_state.sel_quarters
    months = [m for q in qs for m in QUARTER_MONTHS[q] if m in available_months]
    st.session_state.sel_months = sorted(set(months), key=lambda m: MESES_ORDEM.index(m))

# ---------------------------------------------------------------------------
# Filtro de período
# ---------------------------------------------------------------------------
st.markdown(f"""<div class="filtro-box"><b>🗓️ Filtro de Período</b><br>
<span style="font-size:12px;opacity:.85;">Somente meses com dados lançados são exibidos.</span></div>""",
            unsafe_allow_html=True)

qb1, qb2, qb3, qb4 = st.columns(4)
if qb1.button("📅 Todos os meses"):
    st.session_state.sel_months = available_months
    st.session_state.sel_quarters = quarters_available
if qb2.button("📆 Último trimestre"):
    last_q = quarters_available[-1] if quarters_available else None
    if last_q:
        st.session_state.sel_quarters = [last_q]
        st.session_state.sel_months = [m for m in QUARTER_MONTHS[last_q] if m in available_months]
if qb3.button("🗓️ Último mês"):
    if available_months:
        st.session_state.sel_months = [available_months[-1]]
        st.session_state.sel_quarters = [QUARTER_OF[available_months[-1]]]
if qb4.button("🧹 Limpar meses"):
    st.session_state.sel_months = []
    st.session_state.sel_quarters = []

fcol1, fcol2 = st.columns(2)
with fcol1:
    st.multiselect("Trimestre", quarters_available, key="sel_quarters", on_change=sync_quarter_to_months)
with fcol2:
    selected_months = st.multiselect("Meses", available_months, key="sel_months")

if not st.session_state.sel_months:
    st.warning("Selecione ao menos um mês (ou um trimestre) para continuar.")
    st.stop()
selected_months = st.session_state.sel_months

result = aggregate(parsed_list, selected_files, selected_months)
combined, combined_tx, combined_desp, (partner_share, maldivas_share) = result

def fmt_r(v):
    v = 0 if v is None or (isinstance(v, float) and np.isnan(v)) else v
    if mil_mode:
        v = v / 1000
    sign = "-" if v < 0 else ""
    txt = f"{abs(v):,.0f}".replace(",", ".")
    return f"{sign}R$ {txt}{' mil' if mil_mode else ''}"

def fmt_pct(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "-"
    return f"{v:.1%}"

# ---------------------------------------------------------------------------
# KPIs
# ---------------------------------------------------------------------------
receita_mensal = combined["ReceitaDireta"] + combined["ReceitaPortal"]
receita_total = receita_mensal.sum()
custos_totais_mensal = combined["Impostos"] + combined["Custo"] + combined["RebateAAI"] - combined["CoCorretagem"]
custos_totais = custos_totais_mensal.sum()
margem_contribuicao = combined["MargemContribuicao"].sum()
despesas_totais = combined["Despesas"].sum()
resultado_operacional = combined["ResultadoOperacional"].sum()
margem_lucro = (resultado_operacional / receita_total) if receita_total else 0
status = "LUCRO" if resultado_operacional >= 0 else "DÉFICIT"
periodo_label = f"{selected_months[0]} a {selected_months[-1]} 2026" if len(selected_months) > 1 else f"{selected_months[0]} 2026"

delta_receita = month_delta(receita_mensal)
delta_custos = month_delta(custos_totais_mensal)
delta_margem_c = month_delta(combined["MargemContribuicao"])
delta_despesas = month_delta(combined["Despesas"])
delta_resultado = month_delta(combined["ResultadoOperacional"])

# ---------------------------------------------------------------------------
# Header
# ---------------------------------------------------------------------------
logo_html = ""
if st.session_state.get("logo_bytes"):
    b64 = base64.b64encode(st.session_state["logo_bytes"]).decode()
    logo_html = f'<img src="data:image/png;base64,{b64}" style="height:52px;border-radius:10px;margin-right:16px;">'

st.markdown(f"""
<div class="hero" style="display:flex;align-items:center;">
  {logo_html}
  <div>
    <h1>📊 Dashboard Financeiro Premium</h1>
    <p>{" + ".join(selected_files)} · Período: {periodo_label}</p>
    <p>Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}</p>
  </div>
</div>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Insights compartilhados — usados tanto na aba Storytelling quanto na aba
# Apresentação Executiva, para as duas falarem exatamente os mesmos números.
# ---------------------------------------------------------------------------
meses_com_dados = list(combined.index)
n_meses = len(meses_com_dados)

crescimento_total = None
if n_meses >= 2 and receita_mensal.iloc[0]:
    crescimento_total = (receita_mensal.iloc[-1] - receita_mensal.iloc[0]) / abs(receita_mensal.iloc[0])
var_mensal = receita_mensal.pct_change().dropna()
meses_alta = int((var_mensal > 0).sum())
meses_baixa = int((var_mensal < 0).sum())
media_var_mensal = var_mensal.mean() if len(var_mensal) else None
cv_receita = (receita_mensal.std() / receita_mensal.mean()) if receita_mensal.mean() else None
best_month_name = receita_mensal.idxmax()
worst_month_name = receita_mensal.idxmin()

despesas_ratio = (despesas_totais / receita_total) if receita_total else 0
custos_ratio = (custos_totais / receita_total) if receita_total else 0
margem_c_ratio = (margem_contribuicao / receita_total) if receita_total else 0
direta_share = (combined["ReceitaDireta"].sum() / receita_total) if receita_total else 0
portal_share = (combined["ReceitaPortal"].sum() / receita_total) if receita_total else 0

margem_mensal = (combined["ResultadoOperacional"] /
                  (combined["ReceitaDireta"] + combined["ReceitaPortal"]).replace(0, np.nan))
margem_trend_delta = None
if n_meses >= 4:
    metade = n_meses // 2
    margem_1a_metade = margem_mensal.iloc[:metade].mean()
    margem_2a_metade = margem_mensal.iloc[metade:].mean()
    if pd.notna(margem_1a_metade) and pd.notna(margem_2a_metade):
        margem_trend_delta = margem_2a_metade - margem_1a_metade

# concentração de receita por dimensão (depende de haver transações detalhadas)
concentracoes = []
if combined_tx is not None:
    total_tx = combined_tx["Valor"].sum()
    for dim in ["Seguradora", "Produto", "Cliente", "Originador"]:
        if total_tx and (combined_tx[dim] != "Não informado").any():
            top = combined_tx.groupby(dim)["Valor"].sum().sort_values(ascending=False)
            concentracoes.append((dim, top.index[0], top.iloc[0] / total_tx))
max_concentracao = max(concentracoes, key=lambda c: c[2]) if concentracoes else None

tendencia_txt = "sem histórico suficiente para apurar tendência"
if crescimento_total is not None:
    direcao = "crescimento" if crescimento_total >= 0 else "retração"
    tendencia_txt = (f"{direcao} acumulado de {abs(crescimento_total):.1%} entre "
                      f"{meses_com_dados[0]} e {meses_com_dados[-1]}")


tab_dash, tab_story, tab_exec = st.tabs(["📊 Dashboard", "📖 Storytelling", "🎯 Apresentação Executiva"])

with tab_dash:
    # ---------------------------------------------------------------------------
    # Indicadores Principais
    # ---------------------------------------------------------------------------
    st.markdown('<div class="section-title">💰 Indicadores Principais</div>', unsafe_allow_html=True)
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.markdown(f"""<div class="kpi-card"><div class="kpi-label">Faturamento</div>
        <div class="kpi-value">{fmt_r(receita_total)}</div>{delta_html(delta_receita)}</div>""", unsafe_allow_html=True)
    k2.markdown(f"""<div class="kpi-card"><div class="kpi-label">Custos Totais</div>
        <div class="kpi-value">{fmt_r(custos_totais)}</div>{delta_html(delta_custos)}</div>""", unsafe_allow_html=True)
    k3.markdown(f"""<div class="kpi-card"><div class="kpi-label">Margem Contribuição</div>
        <div class="kpi-value">{fmt_r(margem_contribuicao)}</div>{delta_html(delta_margem_c)}</div>""", unsafe_allow_html=True)
    k4.markdown(f"""<div class="kpi-card"><div class="kpi-label">Despesas</div>
        <div class="kpi-value">{fmt_r(despesas_totais)}</div>{delta_html(delta_despesas)}</div>""", unsafe_allow_html=True)
    k5.markdown(f"""<div class="kpi-card"><div class="kpi-label">Resultado Operacional</div>
        <div class="kpi-value">{fmt_r(resultado_operacional)}</div>{delta_html(delta_resultado)}</div>""", unsafe_allow_html=True)

    st.markdown(f"""<div class="legenda-box">
    <b>Faturamento</b>: receita bruta total (Direta + Portal MAAS) · <b>Custos Totais</b>: impostos + custo operacional + rebate AAI − co-corretagem ·
    <b>Margem de Contribuição</b>: faturamento − custos diretos · <b>Despesas</b>: despesas administrativas + folha/terceiros ·
    <b>Resultado Operacional</b>: margem de contribuição − despesas &nbsp;|&nbsp; Margem de lucro: {fmt_pct(margem_lucro)} · Status: <b>{status}</b>
    </div>""", unsafe_allow_html=True)

    # ---------------------------------------------------------------------------
    # Evolução Mensal
    # ---------------------------------------------------------------------------
    st.markdown('<div class="section-title">📈 Evolução Mensal</div>', unsafe_allow_html=True)
    crescimento = receita_mensal.pct_change() * 100

    fig_evol = make_subplots(rows=3, cols=1, shared_xaxes=True, vertical_spacing=0.09,
                              subplot_titles=("Receita Bruta", "Crescimento (%)", "Resultado Operacional"))
    fig_evol.add_bar(x=combined.index, y=receita_mensal, name="Receita Bruta", marker_color=CH_BLUE, row=1, col=1,
                      marker_line_width=0, opacity=0.92,
                      text=[fmt_r(v) for v in receita_mensal], textposition="outside", textfont=dict(size=10),
                      hovertemplate="<b>%{x}</b><br>Receita Bruta: R$ %{y:,.0f}<extra></extra>")
    # realce do melhor e do pior mês de receita, para leitura instantânea da tendência
    best_i, worst_i = int(np.argmax(receita_mensal.values)), int(np.argmin(receita_mensal.values))
    fig_evol.add_annotation(x=combined.index[best_i], y=receita_mensal.values[best_i], row=1, col=1,
                             text="🏆 melhor mês", showarrow=True, arrowhead=2, arrowcolor=CH_SUCCESS,
                             font=dict(color=CH_SUCCESS, size=10), ay=-28)
    if worst_i != best_i:
        fig_evol.add_annotation(x=combined.index[worst_i], y=receita_mensal.values[worst_i], row=1, col=1,
                                 text="mín.", showarrow=True, arrowhead=2, arrowcolor=CH_DANGER_LIGHT,
                                 font=dict(color=CH_DANGER, size=10), ay=24)
    fig_evol.add_trace(go.Scatter(x=combined.index, y=crescimento, mode="lines+markers+text", name="Crescimento %",
                                   line=dict(color=CH_GOLD, width=3, shape="spline", smoothing=0.4),
                                   marker=dict(size=9, line=dict(width=2, color="rgba(10,22,40,.55)")),
                                   text=[f"{v:+.0f}%" if pd.notna(v) else "" for v in crescimento],
                                   textposition="top center", textfont=dict(size=9, color=CH_GOLD),
                                   fill="tozeroy", fillcolor="rgba(254,202,87,.12)",
                                   hovertemplate="<b>%{x}</b><br>Crescimento: %{y:.1f}%<extra></extra>"), row=2, col=1)
    fig_evol.add_hline(y=0, line_dash="dot", line_color="rgba(140,150,160,.6)", row=2, col=1)
    cores_resultado = [CH_SUCCESS if v >= 0 else CH_DANGER for v in combined["ResultadoOperacional"]]
    fig_evol.add_bar(x=combined.index, y=combined["ResultadoOperacional"], name="Resultado Operacional",
                      marker_color=cores_resultado, row=3, col=1, marker_line_width=0, opacity=0.92,
                      text=[fmt_r(v) for v in combined["ResultadoOperacional"]], textposition="outside",
                      textfont=dict(size=10),
                      hovertemplate="<b>%{x}</b><br>Resultado Operacional: R$ %{y:,.0f}<extra></extra>")
    fig_evol.add_hline(y=0, line_dash="dot", line_color="rgba(140,150,160,.6)", row=3, col=1)
    fig_evol.update_layout(showlegend=False, bargap=0.35)
    fig_evol.update_annotations(font_size=12)
    st.plotly_chart(style_fig(fig_evol, height=680, hovermode="x unified"), use_container_width=True, config=PLOTLY_CONFIG)

    tab_r = pd.DataFrame({"Mês": combined.index, "Receita Bruta": receita_mensal.values,
                           "Crescimento": ["-"] + [f"{v:+.1f}%" for v in crescimento.values[1:]],
                           "Resultado Operacional": combined["ResultadoOperacional"].values})
    st.dataframe(tab_r.style.format({"Receita Bruta": fmt_r, "Resultado Operacional": fmt_r}),
                 hide_index=True, use_container_width=True)

    # ---------------------------------------------------------------------------
    # Ranking Seguradoras
    # ---------------------------------------------------------------------------
    st.markdown('<div class="section-title">🏢 Ranking — Seguradoras</div>', unsafe_allow_html=True)
    if combined_tx is not None and (combined_tx["Seguradora"] != "Não informado").any():
        rank_seg = combined_tx.groupby("Seguradora")["Valor"].sum().sort_values(ascending=False).head(15)
        total_seg = combined_tx["Valor"].sum()
        fig_seg = px.bar(rank_seg[::-1], orientation="h", labels={"value": "R$", "Seguradora": ""},
                          color=rank_seg[::-1].values, color_continuous_scale=[CH_BLUE_LIGHT, CH_BLUE])
        fig_seg.update_layout(showlegend=False, coloraxis_showscale=False,
                              height=max(320, 34 * len(rank_seg)))
        fig_seg.update_traces(
            marker_line_width=0,
            text=[fmt_r(v) for v in rank_seg[::-1].values], textposition="outside", textfont=dict(size=10.5),
            customdata=[v / total_seg if total_seg else 0 for v in rank_seg[::-1].values],
            hovertemplate="<b>%{y}</b><br>R$ %{x:,.0f} · %{customdata:.1%} do total<extra></extra>")
        st.plotly_chart(finish_hbar(fig_seg, rank_seg.values, height=max(320, 34 * len(rank_seg))),
                         use_container_width=True, config=PLOTLY_CONFIG)
    else:
        empty_state("🏢", "Coluna de Seguradora não encontrada na aba de receitas — ranking indisponível.")

    # ---------------------------------------------------------------------------
    # Distribuição de Resultados — Sócios
    # ---------------------------------------------------------------------------
    st.markdown('<div class="section-title">🤝 Distribuição de Resultados — Sócios</div>', unsafe_allow_html=True)
    socio_partner_total = combined["SocioPartner"].sum()
    socio_maldivas_total = combined["SocioMaldivas"].sum()
    denom = socio_partner_total + socio_maldivas_total
    partner_pct = (socio_partner_total / denom) if denom else partner_share
    maldivas_pct = 1 - partner_pct

    s1, s2 = st.columns(2)
    s1.markdown(f"""<div class="socio-card" style="background:linear-gradient(135deg,{C_NAVY},{C_NAVY2});">
    <b>👔 Sócio Partner ({partner_pct:.0%})</b><h2 style="margin:6px 0;">{fmt_r(socio_partner_total)}</h2>
    <span style="font-size:12px;opacity:.9;">Participação majoritária no resultado</span></div>""", unsafe_allow_html=True)
    s2.markdown(f"""<div class="socio-card" style="background:linear-gradient(135deg,{C_TEAL},#0B4A63);">
    <b>🏝️ Sócio Maldivas ({maldivas_pct:.0%})</b><h2 style="margin:6px 0;">{fmt_r(socio_maldivas_total)}</h2>
    <span style="font-size:12px;opacity:.9;">Participação operacional</span></div>""", unsafe_allow_html=True)

    fig_soc = go.Figure()
    fig_soc.add_bar(x=combined.index, y=combined["SocioPartner"], name="Sócio Partner", marker_color=CH_PURPLE,
                     marker_line_width=0, opacity=0.92,
                     hovertemplate="<b>%{x}</b><br>Sócio Partner: R$ %{y:,.0f}<extra></extra>")
    fig_soc.add_bar(x=combined.index, y=combined["SocioMaldivas"], name="Sócio Maldivas", marker_color=CH_BLUE,
                     marker_line_width=0, opacity=0.92,
                     hovertemplate="<b>%{x}</b><br>Sócio Maldivas: R$ %{y:,.0f}<extra></extra>")
    fig_soc.add_trace(go.Scatter(x=combined.index, y=combined["ResultadoOperacional"], name="Resultado Total",
                                  mode="lines+markers", line=dict(color=CH_GOLD, width=3, shape="spline", smoothing=0.4),
                                  marker=dict(size=9, line=dict(width=2, color="rgba(10,22,40,.55)")),
                                  hovertemplate="<b>%{x}</b><br>Resultado Total: R$ %{y:,.0f}<extra></extra>"))
    fig_soc.add_hline(y=0, line_dash="dot", line_color="rgba(140,150,160,.6)")
    fig_soc.update_layout(barmode="group", bargap=0.25, bargroupgap=0.08,
                           legend=dict(orientation="h", y=-0.2))
    st.plotly_chart(style_fig(fig_soc, height=420, hovermode="x unified"), use_container_width=True, config=PLOTLY_CONFIG)

    bg_total = "rgba(0,212,170,.14)" if resultado_operacional >= 0 else "rgba(255,107,107,.14)"
    color_total = "#00916e" if resultado_operacional >= 0 else "#d9364a"
    st.markdown(f"""<div class="total-box" style="background:{bg_total};color:{color_total};">
    Resultado total do período: {fmt_r(resultado_operacional)}</div>""", unsafe_allow_html=True)

    quarter_series = combined.groupby([QUARTER_OF[m] for m in combined.index])["ValorPagarMaldivas"].sum()
    quarter_series = quarter_series.reindex([q for q in quarters_available if q in
                                              {QUARTER_OF[m] for m in combined.index}])
    if len(quarter_series):
        st.markdown("**Valor a Receber (Maldivas) por Trimestre**")
        qcols = st.columns(len(quarter_series))
        for col, q in zip(qcols, quarter_series.index):
            col.markdown(f"""<div class="kpi-card"><div class="kpi-label">{q}</div>
            <div class="kpi-value" style="font-size:18px;">{fmt_r(quarter_series[q])}</div></div>""", unsafe_allow_html=True)

    # ---------------------------------------------------------------------------
    # Análise por Produto
    # ---------------------------------------------------------------------------
    st.markdown('<div class="section-title">🧩 Análise por Produto</div>', unsafe_allow_html=True)
    if combined_tx is not None and (combined_tx["Produto"] != "Não informado").any():
        p1, p2 = st.columns([1, 1])
        with p1:
            fig_sun = px.sunburst(combined_tx, path=["Seguradora", "Produto"], values="Valor",
                                   color_discrete_sequence=CH_QUALITATIVE)
            fig_sun.update_traces(hovertemplate="<b>%{label}</b><br>R$ %{value:,.0f} · %{percentParent:.1%} do ramo<extra></extra>",
                                   textinfo="label+percent parent", insidetextorientation="radial",
                                   marker=dict(line=dict(color="rgba(255,255,255,.35)", width=1.5)))
            fig_sun.update_layout(height=440)
            st.plotly_chart(style_fig(fig_sun), use_container_width=True, config=PLOTLY_CONFIG)
        with p2:
            rank_prod = combined_tx.groupby("Produto")["Valor"].sum().sort_values(ascending=False).head(15)
            total_prod = combined_tx["Valor"].sum()
            fig_prod = px.bar(rank_prod[::-1], orientation="h", labels={"value": "R$", "Produto": ""},
                               color=rank_prod[::-1].values, color_continuous_scale=[CH_BLUE_LIGHT, CH_BLUE])
            fig_prod.update_layout(showlegend=False, coloraxis_showscale=False, height=440)
            fig_prod.update_traces(
                marker_line_width=0,
                text=[fmt_r(v) for v in rank_prod[::-1].values], textposition="outside", textfont=dict(size=10.5),
                customdata=[v / total_prod if total_prod else 0 for v in rank_prod[::-1].values],
                hovertemplate="<b>%{y}</b><br>R$ %{x:,.0f} · %{customdata:.1%} do total<extra></extra>")
            st.plotly_chart(finish_hbar(fig_prod, rank_prod.values, height=440),
                             use_container_width=True, config=PLOTLY_CONFIG)
    else:
        empty_state("🧩", "Coluna de Produto não encontrada na aba de receitas — análise indisponível.")

    # ---------------------------------------------------------------------------
    # Ranking Originadores
    # ---------------------------------------------------------------------------
    st.markdown('<div class="section-title">👥 Ranking — Originadores</div>', unsafe_allow_html=True)
    ranking_df = None
    if combined_tx is not None and (combined_tx["Originador"] != "Não informado").any():
        ranking_full = (combined_tx.groupby("Originador")
                         .agg(Valor=("Valor", "sum"), Operacoes=("Valor", "count"))
                         .assign(TicketMedio=lambda d: d["Valor"] / d["Operacoes"])
                         .sort_values("Valor", ascending=False))
        ranking_df = ranking_full.head(3).reset_index()
        o1, o2 = st.columns([1, 1])
        with o1:
            top_n = ranking_full.head(5)
            outros = ranking_full["Valor"].iloc[5:].sum()
            labels = list(top_n.index) + (["Outros"] if outros > 0 else [])
            values = list(top_n["Valor"]) + ([outros] if outros > 0 else [])
            pulls = [0.06] + [0] * (len(labels) - 1)
            fig_don = go.Figure(go.Pie(
                labels=labels, values=values, hole=.62, pull=pulls, sort=False,
                marker=dict(colors=CH_QUALITATIVE, line=dict(color="rgba(255,255,255,.4)", width=2)),
                textinfo="percent", textfont=dict(size=12, color="white"),
                hovertemplate="<b>%{label}</b><br>R$ %{value:,.0f} · %{percent}<extra></extra>"))
            fig_don.add_annotation(text=f"<b>{fmt_r(sum(values))}</b><br><span style='font-size:11px;'>total</span>",
                                    showarrow=False, font=dict(size=15, color="#8a99ab"))
            fig_don.update_layout(title="Participação por Originador")
            st.plotly_chart(style_fig(fig_don, height=420), use_container_width=True, config=PLOTLY_CONFIG)
        with o2:
            medals, colors_m = ["🥇", "🥈", "🥉"], [C_GOLD_MEDAL, C_SILVER, C_BRONZE]
            for i, row in ranking_df.iterrows():
                st.markdown(f"""<div class="rank-card" style="border-left:6px solid {colors_m[i]};">
                <div class="rank-medal">{medals[i]}</div>
                <div><b>{row['Originador']}</b><br>
                <span style="font-size:20px;font-weight:800;color:{C_NAVY2};">{fmt_r(row['Valor'])}</span><br>
                <span style="font-size:12px;color:#8c8c8c;">{int(row['Operacoes'])} operações | Ticket médio: {fmt_r(row['TicketMedio'])}</span>
                </div></div>""", unsafe_allow_html=True)
    else:
        empty_state("👥", "Coluna de Originador não encontrada — ranking indisponível.")

    # ---------------------------------------------------------------------------
    # Ranking Clientes
    # ---------------------------------------------------------------------------
    st.markdown('<div class="section-title">🧑‍💼 Ranking — Clientes</div>', unsafe_allow_html=True)
    if combined_tx is not None and (combined_tx["Cliente"] != "Não informado").any():
        rank_cli = combined_tx.groupby("Cliente")["Valor"].sum().sort_values(ascending=False).head(15)
        total_cli = combined_tx["Valor"].sum()
        fig_cli = px.bar(rank_cli[::-1], orientation="h", labels={"value": "R$", "Cliente": ""},
                          color=rank_cli[::-1].values, color_continuous_scale=[CH_BLUE_LIGHT, CH_BLUE])
        fig_cli.update_layout(showlegend=False, coloraxis_showscale=False,
                              height=max(320, 34 * len(rank_cli)))
        fig_cli.update_traces(
            marker_line_width=0,
            text=[fmt_r(v) for v in rank_cli[::-1].values], textposition="outside", textfont=dict(size=10.5),
            customdata=[v / total_cli if total_cli else 0 for v in rank_cli[::-1].values],
            hovertemplate="<b>%{y}</b><br>R$ %{x:,.0f} · %{customdata:.1%} do total<extra></extra>")
        st.plotly_chart(finish_hbar(fig_cli, rank_cli.values, height=max(320, 34 * len(rank_cli))),
                         use_container_width=True, config=PLOTLY_CONFIG)
    else:
        empty_state("🧑‍💼", "Coluna de Cliente não encontrada na aba de receitas — ranking indisponível.")

    # ---------------------------------------------------------------------------
    # Ranking Despesas
    # ---------------------------------------------------------------------------
    st.markdown('<div class="section-title">💸 Ranking — Despesas por Categoria</div>', unsafe_allow_html=True)
    if combined_desp is not None:
        rank_desp = combined_desp.groupby("Categoria")["Valor"].sum().sort_values(ascending=False).head(10)
        total_desp = combined_desp["Valor"].sum()
        fig_desp = px.bar(rank_desp[::-1], orientation="h", labels={"value": "R$", "Categoria": ""},
                           color=rank_desp[::-1].values, color_continuous_scale=[CH_DANGER_LIGHT, CH_DANGER])
        fig_desp.update_layout(showlegend=False, coloraxis_showscale=False,
                               height=max(320, 34 * len(rank_desp)))
        fig_desp.update_traces(
            marker_line_width=0,
            text=[fmt_r(v) for v in rank_desp[::-1].values], textposition="outside", textfont=dict(size=10.5),
            customdata=[v / total_desp if total_desp else 0 for v in rank_desp[::-1].values],
            hovertemplate="<b>%{y}</b><br>R$ %{x:,.0f} · %{customdata:.1%} do total<extra></extra>")
        st.plotly_chart(finish_hbar(fig_desp, rank_desp.values, height=max(320, 34 * len(rank_desp))),
                         use_container_width=True, config=PLOTLY_CONFIG)
    else:
        empty_state("💸", "Aba 'DESPESAS' não encontrada (ou sem colunas Categoria/Valor) — ranking indisponível.")

    # ---------------------------------------------------------------------------
    # Resumo Executivo
    # ---------------------------------------------------------------------------
    resumo_linhas = [
        ("FATURAMENTO BRUTO", receita_total, True),
        ("Produção Direta", combined["ReceitaDireta"].sum(), False),
        ("Portal MAAS", combined["ReceitaPortal"].sum(), False),
        ("Impostos Diretos", combined["Impostos"].sum(), False),
        ("Custo Operacional (D.A)", combined["Custo"].sum(), False),
        ("Co-Corretagem", combined["CoCorretagem"].sum(), False),
        ("Rebate AAI", combined["RebateAAI"].sum(), False),
        ("CUSTOS TOTAIS", custos_totais, True),
        ("(=) MARGEM DE CONTRIBUIÇÃO", margem_contribuicao, True),
        ("DESPESAS TOTAIS", despesas_totais, True),
        ("Folha + Terceiros", combined["Folha"].sum(), False),
        ("RESULTADO OPERACIONAL", resultado_operacional, True),
    ]
    st.markdown('<div class="section-title">📋 Resumo Executivo</div>', unsafe_allow_html=True)
    resumo_html = "<table style='width:100%;border-collapse:collapse;'>"
    for label, val, bold in resumo_linhas:
        w = "700" if bold else "400"
        bg = "#e8f0fe" if bold else "#ffffff"
        resumo_html += f"""<tr style="background:{bg};">
        <td style="padding:9px 12px;font-weight:{w};color:{C_TEXT_DARK};border-bottom:1px solid #eee;">{label}</td>
        <td style="padding:9px 12px;text-align:right;font-weight:{w};color:{C_TEXT_DARK};border-bottom:1px solid #eee;">{fmt_r(val)}</td></tr>"""
    resumo_html += "</table>"
    st.markdown(resumo_html, unsafe_allow_html=True)

    # ---------------------------------------------------------------------------
    # PDF — ReportLab
    # ---------------------------------------------------------------------------
    def build_pdf():
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.2 * cm, bottomMargin=1.6 * cm,
                                 leftMargin=1.5 * cm, rightMargin=1.5 * cm)
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle("CoverTitle", fontSize=26, textColor=colors.white, fontName="Helvetica-Bold",
                                   alignment=TA_CENTER, leading=30))
        styles.add(ParagraphStyle("CoverSub", fontSize=13, textColor=colors.white, fontName="Helvetica",
                                   alignment=TA_CENTER, leading=18))
        styles.add(ParagraphStyle("SectionTitle", fontSize=15, textColor=HexColor(C_NAVY2), fontName="Helvetica-Bold",
                                   spaceBefore=10, spaceAfter=5, keepWithNext=1))
        styles.add(ParagraphStyle("Body", fontSize=10, textColor=HexColor(C_TEXT_DARK), fontName="Helvetica", leading=14))

        story = []

        # A capa usa um retângulo de fundo desenhado direto no canvas (ver draw_cover_bg
        # em onFirstPage) em vez de uma Table dimensionada pelo conteúdo — isso garante que
        # a cor cubra a página inteira, sem a grande faixa branca que sobrava abaixo dela.
        story.append(Spacer(1, 9.5 * cm))
        story.append(Paragraph("Dashboard Financeiro Premium", styles["CoverTitle"]))
        story.append(Spacer(1, 0.5 * cm))
        story.append(Paragraph(" + ".join(selected_files), styles["CoverSub"]))
        story.append(Paragraph(f"Período analisado: {periodo_label}", styles["CoverSub"]))
        story.append(Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", styles["CoverSub"]))
        story.append(PageBreak())

        story.append(Paragraph("Sumário", styles["SectionTitle"]))
        secoes = ["Indicadores Principais", "Evolução Mensal", "Ranking de Seguradoras",
                  "Distribuição de Resultados (Sócios)", "Ranking de Originadores", "Ranking de Clientes",
                  "Ranking de Despesas", "Resumo Executivo"]
        for i, s in enumerate(secoes, 1):
            story.append(Paragraph(f"{i}. {s}", styles["Body"]))
        story.append(Spacer(1, 0.3 * cm))

        story.append(Paragraph("Indicadores Principais", styles["SectionTitle"]))
        kpi_data = [["Faturamento", "Custos Totais", "Margem Contribuição", "Despesas", "Resultado Operacional"],
                    [fmt_r(receita_total), fmt_r(custos_totais), fmt_r(margem_contribuicao),
                     fmt_r(despesas_totais), fmt_r(resultado_operacional)]]
        kpi_table = Table(kpi_data, colWidths=[3.4 * cm] * 5)
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HexColor(C_NAVY2)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BACKGROUND", (0, 1), (-1, 1), HexColor("#e8f0fe")),
            ("TEXTCOLOR", (0, 1), (-1, 1), HexColor(C_TEXT_DARK)),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(kpi_table)
        story.append(Paragraph(f"Margem de lucro: {fmt_pct(margem_lucro)} · Status: {status}", styles["Body"]))
        story.append(Spacer(1, 0.2 * cm))

        story.append(Paragraph("Evolução Mensal", styles["SectionTitle"]))
        drawing = Drawing(420, 155)
        lc = HorizontalLineChart()
        lc.x, lc.y, lc.width, lc.height = 40, 15, 360, 120
        lc.data = [list(receita_mensal.values), list(combined["ResultadoOperacional"].values)]
        lc.categoryAxis.categoryNames = list(combined.index)
        lc.lines[0].strokeColor = HexColor(C_TEAL)
        lc.lines[1].strokeColor = HexColor(CH_SUCCESS)
        lc.lines[0].strokeWidth = 2
        lc.lines[1].strokeWidth = 2
        drawing.add(lc)
        story.append(drawing)
        story.append(Paragraph("Azul: Receita Bruta · Verde: Resultado Operacional", styles["Body"]))

        mensal_rows = [["Mês", "Receita Bruta", "Resultado Operacional"]]
        for m in combined.index:
            mensal_rows.append([m, fmt_r(receita_mensal[m]), fmt_r(combined["ResultadoOperacional"][m])])
        t_mensal = Table(mensal_rows, colWidths=[3 * cm, 6 * cm, 6 * cm])
        t_mensal.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HexColor(C_NAVY2)), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("TEXTCOLOR", (0, 1), (-1, -1), HexColor(C_TEXT_DARK)),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")), ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(Spacer(1, 0.2 * cm))
        story.append(t_mensal)

        def add_ranking_table(title, df_or_series, value_label="Valor"):
            story.append(Paragraph(title, styles["SectionTitle"]))
            if df_or_series is None or len(df_or_series) == 0:
                story.append(Paragraph("Dados não disponíveis nesta planilha.", styles["Body"]))
                return
            rows = [[df_or_series.index.name or "Item", value_label]]
            for idx, val in df_or_series.items():
                rows.append([str(idx), fmt_r(val)])
            t = Table(rows, colWidths=[10 * cm, 5 * cm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), HexColor(C_NAVY2)), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("TEXTCOLOR", (0, 1), (-1, -1), HexColor(C_TEXT_DARK)),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")), ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t)

        story.append(Spacer(1, 0.3 * cm))
        if combined_tx is not None and (combined_tx["Seguradora"] != "Não informado").any():
            add_ranking_table("Ranking — Seguradoras (Top 15)",
                               combined_tx.groupby("Seguradora")["Valor"].sum().sort_values(ascending=False).head(15))
        else:
            add_ranking_table("Ranking — Seguradoras", None)

        story.append(Spacer(1, 0.25 * cm))
        story.append(Paragraph("Distribuição de Resultados — Sócios", styles["SectionTitle"]))
        socio_rows = [["Sócio", "Participação", "Valor"],
                      ["Partner", f"{partner_pct:.0%}", fmt_r(socio_partner_total)],
                      ["Maldivas", f"{maldivas_pct:.0%}", fmt_r(socio_maldivas_total)]]
        t_socio = Table(socio_rows, colWidths=[5 * cm, 4 * cm, 6 * cm])
        t_socio.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HexColor(C_NAVY2)), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("TEXTCOLOR", (0, 1), (-1, -1), HexColor(C_TEXT_DARK)),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")), ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t_socio)

        story.append(Spacer(1, 0.3 * cm))
        if ranking_df is not None:
            add_ranking_table("Ranking — Originadores (Top 3)", ranking_df.set_index("Originador")["Valor"])
        else:
            add_ranking_table("Ranking — Originadores", None)

        story.append(Spacer(1, 0.25 * cm))
        if combined_tx is not None and (combined_tx["Cliente"] != "Não informado").any():
            add_ranking_table("Ranking — Clientes (Top 15)",
                               combined_tx.groupby("Cliente")["Valor"].sum().sort_values(ascending=False).head(15))
        else:
            add_ranking_table("Ranking — Clientes", None)

        story.append(Spacer(1, 0.25 * cm))
        if combined_desp is not None:
            add_ranking_table("Ranking — Despesas (Top 10)",
                               combined_desp.groupby("Categoria")["Valor"].sum().sort_values(ascending=False).head(10))
        else:
            add_ranking_table("Ranking — Despesas", None)

        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph("Resumo Executivo", styles["SectionTitle"]))
        resumo_rows = [["Indicador", "Valor"]] + [[label, fmt_r(val)] for label, val, _ in resumo_linhas]
        t_resumo = Table(resumo_rows, colWidths=[11 * cm, 5 * cm])
        bold_rows = [i for i, (_, _, bold) in enumerate(resumo_linhas, start=1) if bold]
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), HexColor(C_NAVY2)), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("TEXTCOLOR", (0, 1), (-1, -1), HexColor(C_TEXT_DARK)),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")), ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        for r in bold_rows:
            style_cmds.append(("FONTNAME", (0, r), (-1, r), "Helvetica-Bold"))
            style_cmds.append(("BACKGROUND", (0, r), (-1, r), HexColor("#e8f0fe")))
        t_resumo.setStyle(TableStyle(style_cmds))
        story.append(t_resumo)

        def add_page_number(canvas, doc_):
            canvas.saveState()
            canvas.setFont("Helvetica", 8)
            canvas.setFillColor(HexColor("#888888"))
            canvas.drawCentredString(A4[0] / 2, 1.0 * cm, f"Página {doc_.page} — Dashboard Financeiro Premium")
            canvas.restoreState()

        def draw_cover_page(canvas, doc_):
            canvas.saveState()
            canvas.setFillColor(HexColor(C_NAVY))
            canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
            canvas.setFont("Helvetica", 8)
            canvas.setFillColor(colors.white)
            canvas.drawCentredString(A4[0] / 2, 1.0 * cm, f"Página {doc_.page} — Dashboard Financeiro Premium")
            canvas.restoreState()

        doc.build(story, onFirstPage=draw_cover_page, onLaterPages=add_page_number)
        buf.seek(0)
        return buf.getvalue()

    st.markdown("---")
    st.markdown('<div class="section-title">📤 Exportar</div>', unsafe_allow_html=True)
    pdf_bytes = build_pdf()
    st.download_button("⬇️ Baixar Dashboard em PDF", data=pdf_bytes,
                        file_name=f"dashboard_financeiro_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                        mime="application/pdf")

# ---------------------------------------------------------------------------
# Storytelling — narrativa executiva construída em cima das mesmas métricas
# já calculadas acima (nenhuma fonte de dado nova; apenas leitura interpretativa).
# ---------------------------------------------------------------------------
with tab_story:
    # ---- Hero: leitura executiva em uma frase ----
    st.markdown(f"""<div class="story-hero">
    <div class="story-kicker">📖 Leitura Executiva · {periodo_label}</div>
    <h2>{'📈' if resultado_operacional >= 0 else '📉'} A operação fechou o período com {fmt_r(resultado_operacional)}
    de resultado operacional ({status.lower()}), sobre {fmt_r(receita_total)} faturados — {tendencia_txt}.</h2>
    <p>Esta aba traduz os números do dashboard em uma narrativa: de onde veio o faturamento, como ele se comportou
    mês a mês, quanto foi consumido por custos e despesas, e onde estão as maiores concentrações de risco e
    oportunidade da operação.</p>
    </div>""", unsafe_allow_html=True)

    # ---- Trajetória de Receita ----
    st.markdown('<div class="story-section-title">📈 Trajetória de Receita</div>', unsafe_allow_html=True)
    st.markdown(f"""<p class="story-p">Ao longo de {n_meses} mês(es) com dados lançados, o faturamento médio mensal
    foi de {fmt_r(receita_mensal.mean())}, com {meses_alta} mês(es) de alta e {meses_baixa} mês(es) de queda em
    relação ao mês anterior{f" (variação média de {media_var_mensal:+.1%} ao mês)" if media_var_mensal is not None else ""}.
    O melhor resultado de faturamento foi em <b>{best_month_name}</b> ({fmt_r(receita_mensal[best_month_name])}) e o
    mais fraco em <b>{worst_month_name}</b> ({fmt_r(receita_mensal[worst_month_name])}).</p>""", unsafe_allow_html=True)

    if cv_receita is not None:
        vol_label = "baixa" if cv_receita < 0.15 else ("moderada" if cv_receita < 0.35 else "alta")
        vol_class = "good" if vol_label == "baixa" else ("warn" if vol_label == "moderada" else "bad")
        vol_note = f"coeficiente de variação de {cv_receita:.0%}"
    else:
        vol_label, vol_class, vol_note = "n/d", "", "histórico insuficiente"

    g1, g2, g3, g4 = st.columns(4)
    g1.markdown(f"""<div class="story-card {'good' if (crescimento_total or 0) >= 0 else 'bad'}">
    <div class="story-card-title">Variação no Período</div>
    <div class="story-card-value">{f"{crescimento_total:+.1%}" if crescimento_total is not None else "-"}</div>
    <div class="story-card-note">{meses_com_dados[0]} → {meses_com_dados[-1]}</div></div>""", unsafe_allow_html=True)
    g2.markdown(f"""<div class="story-card good">
    <div class="story-card-title">Melhor Mês</div>
    <div class="story-card-value">{best_month_name}</div>
    <div class="story-card-note">{fmt_r(receita_mensal[best_month_name])}</div></div>""", unsafe_allow_html=True)
    g3.markdown(f"""<div class="story-card {'bad' if worst_month_name != best_month_name else 'good'}">
    <div class="story-card-title">Mês Mais Fraco</div>
    <div class="story-card-value">{worst_month_name}</div>
    <div class="story-card-note">{fmt_r(receita_mensal[worst_month_name])}</div></div>""", unsafe_allow_html=True)
    g4.markdown(f"""<div class="story-card {vol_class}">
    <div class="story-card-title">Volatilidade da Receita</div>
    <div class="story-card-value">{vol_label.capitalize()}</div>
    <div class="story-card-note">{vol_note}</div></div>""", unsafe_allow_html=True)

    # ---- Estrutura de Custos e Rentabilidade ----
    st.markdown('<div class="story-section-title">🧮 Estrutura de Custos e Rentabilidade</div>', unsafe_allow_html=True)
    st.markdown(f"""<p class="story-p">De cada R$ 1,00 faturado, {custos_ratio:.0%} foram consumidos por impostos,
    custo operacional e rebates diretos, restando uma margem de contribuição de {margem_c_ratio:.0%}. As despesas
    administrativas e de pessoal absorveram mais {despesas_ratio:.0%} da receita, resultando em margem líquida de
    {margem_lucro:.1%}. A Produção Direta respondeu por {direta_share:.0%} do faturamento e o Portal MAAS por
    {portal_share:.0%}.</p>""", unsafe_allow_html=True)

    if margem_trend_delta is not None and margem_trend_delta > 0.02:
        st.markdown(f"""<div class="callout good"><b>📈 Rentabilidade em melhora:</b> a margem operacional média da
        segunda metade do período ficou {margem_trend_delta:+.1%} p.p. acima da primeira metade — sinal de ganho de
        eficiência ou de mix de receita mais rentável.</div>""", unsafe_allow_html=True)
    elif margem_trend_delta is not None and margem_trend_delta < -0.02:
        st.markdown(f"""<div class="callout warn"><b>📉 Rentabilidade em queda:</b> a margem operacional média caiu
        {abs(margem_trend_delta):.1%} p.p. da primeira para a segunda metade do período — vale investigar se custos
        e despesas cresceram acima da receita.</div>""", unsafe_allow_html=True)
    elif margem_trend_delta is not None:
        st.markdown("""<div class="callout info"><b>➖ Rentabilidade estável:</b> a margem operacional se manteve
        praticamente constante ao longo do período.</div>""", unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(f"""<div class="story-card"><div class="story-card-title">Custos / Receita</div>
    <div class="story-card-value">{custos_ratio:.0%}</div>
    <div class="story-card-note">impostos + custo operacional + rebate − co-corretagem</div></div>""", unsafe_allow_html=True)
    c2.markdown(f"""<div class="story-card"><div class="story-card-title">Despesas / Receita</div>
    <div class="story-card-value">{despesas_ratio:.0%}</div>
    <div class="story-card-note">administrativas + folha/terceiros</div></div>""", unsafe_allow_html=True)
    c3.markdown(f"""<div class="story-card good"><div class="story-card-title">Margem de Contribuição</div>
    <div class="story-card-value">{margem_c_ratio:.0%}</div>
    <div class="story-card-note">{fmt_r(margem_contribuicao)}</div></div>""", unsafe_allow_html=True)
    c4.markdown(f"""<div class="story-card {'good' if margem_lucro >= 0 else 'bad'}">
    <div class="story-card-title">Margem Líquida</div>
    <div class="story-card-value">{margem_lucro:.1%}</div>
    <div class="story-card-note">resultado operacional / faturamento</div></div>""", unsafe_allow_html=True)

    # ---- Concentração e Risco ----
    if concentracoes:
        st.markdown('<div class="story-section-title">🎯 Concentração e Risco</div>', unsafe_allow_html=True)
        st.markdown("""<p class="story-p">Concentração de receita em poucos nomes eleva o risco de dependência: a
        perda de um único cliente, originador, seguradora ou produto pode comprometer parte relevante do
        faturamento. Os números abaixo mostram a fatia do maior nome em cada dimensão.</p>""", unsafe_allow_html=True)
        cc = st.columns(len(concentracoes))
        for col, (dim, nome, share) in zip(cc, concentracoes):
            classe = "bad" if share >= 0.5 else ("warn" if share >= 0.3 else "good")
            col.markdown(f"""<div class="story-card {classe}"><div class="story-card-title">Top {dim}</div>
            <div class="story-card-value">{share:.0%}</div>
            <div class="story-card-note">{nome}</div></div>""", unsafe_allow_html=True)
        if max_concentracao and max_concentracao[2] >= 0.4:
            dim, nome, share = max_concentracao
            st.markdown(f"""<div class="callout {'bad' if share >= 0.5 else 'warn'}"><b>⚠️ Risco de concentração:</b>
            <b>{nome}</b> responde sozinho(a) por {share:.0%} da receita na dimensão <b>{dim}</b>. Recomenda-se um
            plano de diversificação para reduzir a dependência desse único nome.</div>""", unsafe_allow_html=True)
        else:
            st.markdown("""<div class="callout good"><b>✅ Base pulverizada:</b> nenhuma seguradora, produto, cliente
            ou originador concentra isoladamente uma fatia crítica da receita — a carteira está razoavelmente
            distribuída.</div>""", unsafe_allow_html=True)
    else:
        empty_state("🎯", "Sem colunas de transação suficientes para analisar concentração de risco nesta planilha.")

    # ---- Distribuição entre Sócios ----
    st.markdown('<div class="story-section-title">🤝 Distribuição entre Sócios</div>', unsafe_allow_html=True)
    alinhamento = "em linha com" if abs(partner_pct - partner_share) < 0.05 else "com desvio em relação a"
    st.markdown(f"""<p class="story-p">Do resultado operacional total de {fmt_r(resultado_operacional)}, o Sócio
    Partner ficou com {fmt_r(socio_partner_total)} ({partner_pct:.0%}) e o Sócio Maldivas com
    {fmt_r(socio_maldivas_total)} ({maldivas_pct:.0%}), {alinhamento} a divisão contratual de referência
    ({partner_share:.0%} / {maldivas_share:.0%}).</p>""", unsafe_allow_html=True)

    quarter_recv_story = combined.groupby([QUARTER_OF[m] for m in combined.index])["ValorPagarMaldivas"].sum()
    quarter_recv_story = quarter_recv_story.reindex(
        [q for q in quarters_available if q in {QUARTER_OF[m] for m in combined.index}])
    if len(quarter_recv_story):
        melhor_tri = quarter_recv_story.idxmax()
        st.markdown(f"""<p class="story-p">Em bases trimestrais, o trimestre com maior valor a receber pela Maldivas
        foi <b>{melhor_tri}</b> ({fmt_r(quarter_recv_story[melhor_tri])}).</p>""", unsafe_allow_html=True)

    # ---- Despesas em destaque ----
    if combined_desp is not None:
        rank_desp_story = combined_desp.groupby("Categoria")["Valor"].sum().sort_values(ascending=False)
        total_desp_story = rank_desp_story.sum()
        if len(rank_desp_story) and total_desp_story:
            top_cat, top_val = rank_desp_story.index[0], rank_desp_story.iloc[0]
            st.markdown(f"""<div class="callout info"><b>💸 Maior categoria de despesa:</b> <b>{top_cat}</b>
            concentra {fmt_r(top_val)} ({top_val / total_desp_story:.0%} das despesas detalhadas na planilha) —
            a principal alavanca de controle de custo administrativo no período.</div>""", unsafe_allow_html=True)

    # ---- Recomendações Estratégicas ----
    st.markdown('<div class="story-section-title">🧭 Recomendações Estratégicas</div>', unsafe_allow_html=True)
    recos = []
    if status == "DÉFICIT":
        recos.append("Priorizar um plano de recuperação de curto prazo: revisar despesas administrativas e "
                      "renegociar custos operacionais até reverter o déficit operacional.")
    if margem_trend_delta is not None and margem_trend_delta < -0.02:
        recos.append("Investigar a queda de margem entre a primeira e a segunda metade do período — comparar o "
                      "crescimento de custos/despesas frente ao crescimento de receita mês a mês.")
    if max_concentracao and max_concentracao[2] >= 0.4:
        dim, nome, share = max_concentracao
        recos.append(f"Reduzir a dependência de <b>{nome}</b> ({share:.0%} da receita em {dim}) diversificando a "
                      f"base — meta sugerida: nenhum nome isolado acima de 30%.")
    if despesas_ratio > 0.3:
        recos.append(f"Revisar a estrutura de despesas administrativas, que hoje representam {despesas_ratio:.0%} "
                      f"da receita — buscar eficiência em folha e terceiros.")
    if crescimento_total is not None and crescimento_total > 0.1:
        recos.append("Aproveitar o momentum de crescimento para negociar melhores condições comerciais com "
                      "seguradoras e reforçar a originação nos meses de sazonalidade mais forte.")
    if cv_receita is not None and cv_receita >= 0.35:
        recos.append("Suavizar a sazonalidade da receita com ações comerciais recorrentes — a alta volatilidade "
                      "mensal dificulta o planejamento de caixa.")
    if not recos:
        recos.append("Manter a governança atual: os indicadores do período não apontam riscos críticos de "
                      "concentração, rentabilidade ou caixa.")
    recos.append("Repetir esta leitura mensalmente para acompanhar a evolução dos indicadores acima e antecipar "
                  "desvios de rota.")

    reco_html = "".join(
        f'<div class="reco-item"><div class="reco-num">{i}</div><div class="reco-text">{r}</div></div>'
        for i, r in enumerate(recos, 1)
    )
    st.markdown(f'<div class="story-card" style="border-left-color:{C_TEAL};">{reco_html}</div>', unsafe_allow_html=True)

    # ---- Conclusão ----
    st.markdown('<div class="story-section-title">✅ Conclusão</div>', unsafe_allow_html=True)
    foco_risco = ("O principal ponto de atenção é a concentração de receita" if
                  (max_concentracao and max_concentracao[2] >= 0.4) else
                  "A base de receita está razoavelmente distribuída")
    st.markdown(f"""<p class="story-p">Em síntese, o período de {periodo_label} fecha com <b>{status.lower()}</b>
    operacional de {fmt_r(resultado_operacional)} sobre {fmt_r(receita_total)} de faturamento (margem líquida de
    {margem_lucro:.1%}). {foco_risco}, e a estrutura de custos consome {custos_ratio:.0%} do faturamento antes das
    despesas administrativas. As recomendações acima priorizam, nesta ordem, saúde de caixa, rentabilidade e
    diversificação de risco.</p>""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Apresentação Executiva — visão enxuta para CFO/CEO: só os gráficos que
# realmente sustentam uma decisão sobre "como vai a produção". Usa os mesmos
# agregados computados acima (nenhum número novo, mesma fonte de verdade).
# ---------------------------------------------------------------------------
with tab_exec:
    st.markdown(f"""<div class="exec-hero">
    <div class="exec-kicker">🎯 Apresentação Executiva · Para CFO &amp; CEO</div>
    <h1>Como vai a produção — {periodo_label}</h1>
    <p>{" + ".join(selected_files)} · {n_meses} mês(es) analisado(s) · Gerado em
    {datetime.now().strftime('%d/%m/%Y às %H:%M')}</p>
    </div>""", unsafe_allow_html=True)

    e1, e2, e3, e4 = st.columns(4)
    e1.markdown(f"""<div class="exec-tile"><div class="exec-tile-label">Faturamento</div>
    <div class="exec-tile-value">{fmt_r(receita_total)}</div>{delta_html(delta_receita)}</div>""", unsafe_allow_html=True)
    e2.markdown(f"""<div class="exec-tile"><div class="exec-tile-label">Resultado Operacional</div>
    <div class="exec-tile-value">{fmt_r(resultado_operacional)}</div>{delta_html(delta_resultado)}</div>""", unsafe_allow_html=True)
    e3.markdown(f"""<div class="exec-tile"><div class="exec-tile-label">Margem Líquida</div>
    <div class="exec-tile-value">{margem_lucro:.1%}</div>
    <span style="font-size:12px;color:#8a99ab;">status: {status.lower()}</span></div>""", unsafe_allow_html=True)
    e4.markdown(f"""<div class="exec-tile"><div class="exec-tile-label">Variação no Período</div>
    <div class="exec-tile-value">{f"{crescimento_total:+.1%}" if crescimento_total is not None else "-"}</div>
    <span style="font-size:12px;color:#8a99ab;">{meses_com_dados[0]} → {meses_com_dados[-1]}</span></div>""",
                unsafe_allow_html=True)

    # ---- Slide 1 — Evolução da Produção ----
    st.markdown("""<div class="exec-slide">
    <div class="exec-slide-title"><span class="exec-slide-num">1</span>Evolução da Produção</div>""",
                unsafe_allow_html=True)
    tendencia_lbl = "crescimento" if (crescimento_total or 0) >= 0 else "retração"
    st.markdown(f"""<div class="exec-headline">Faturamento saiu de <b>{fmt_r(receita_mensal.iloc[0])}</b> em
    <b>{meses_com_dados[0]}</b> para <b>{fmt_r(receita_mensal.iloc[-1])}</b> em <b>{meses_com_dados[-1]}</b> —
    {tendencia_lbl} acumulado(a) de <b>{f"{crescimento_total:+.1%}" if crescimento_total is not None else "n/d"}</b>,
    com o melhor mês em <b>{best_month_name}</b> e o mais fraco em <b>{worst_month_name}</b>.</div>""",
                unsafe_allow_html=True)

    x_idx = np.arange(len(receita_mensal))
    if len(x_idx) >= 2:
        coef = np.polyfit(x_idx, receita_mensal.values, 1)
        tendencia_linha = np.poly1d(coef)(x_idx)
    else:
        tendencia_linha = receita_mensal.values

    fig_prod_evol = go.Figure()
    fig_prod_evol.add_bar(x=combined.index, y=receita_mensal, name="Receita Bruta", marker_color=CH_BLUE,
                           marker_line_width=0, opacity=0.92,
                           text=[fmt_r(v) for v in receita_mensal], textposition="outside", textfont=dict(size=10),
                           hovertemplate="<b>%{x}</b><br>Receita Bruta: R$ %{y:,.0f}<extra></extra>")
    fig_prod_evol.add_trace(go.Scatter(x=combined.index, y=tendencia_linha, mode="lines", name="Tendência",
                                        line=dict(color=CH_GOLD, width=3, dash="dash"),
                                        hovertemplate="Tendência: R$ %{y:,.0f}<extra></extra>"))
    fig_prod_evol.update_layout(showlegend=True, bargap=0.35)
    st.plotly_chart(style_fig(fig_prod_evol, height=380, hovermode="x unified"),
                     use_container_width=True, config=PLOTLY_CONFIG)
    st.markdown("</div>", unsafe_allow_html=True)

    # ---- Slide 2 — Performance por Originador ----
    st.markdown("""<div class="exec-slide">
    <div class="exec-slide-title"><span class="exec-slide-num">2</span>Performance por Originador</div>""",
                unsafe_allow_html=True)
    if combined_tx is not None and (combined_tx["Originador"] != "Não informado").any():
        rank_ori_exec = combined_tx.groupby("Originador")["Valor"].sum().sort_values(ascending=False)
        total_ori_exec = rank_ori_exec.sum()
        top_ori_nome, top_ori_val = rank_ori_exec.index[0], rank_ori_exec.iloc[0]
        top_ori_share = top_ori_val / total_ori_exec if total_ori_exec else 0
        risco_txt = (f" — concentração alta, vale atenção" if top_ori_share >= 0.4 else "")
        st.markdown(f"""<div class="exec-headline"><b>{top_ori_nome}</b> lidera a produção com
        {fmt_r(top_ori_val)} ({top_ori_share:.0%} do total){risco_txt}. A equipe conta com
        {rank_ori_exec.shape[0]} originador(es) ativo(s) no período.</div>""", unsafe_allow_html=True)
        top8_ori = rank_ori_exec.head(8)
        fig_ori_exec = px.bar(top8_ori[::-1], orientation="h", labels={"value": "R$", "Originador": ""},
                               color=top8_ori[::-1].values, color_continuous_scale=[CH_BLUE_LIGHT, CH_BLUE])
        fig_ori_exec.update_layout(showlegend=False, coloraxis_showscale=False, height=max(280, 34 * len(top8_ori)))
        fig_ori_exec.update_traces(
            marker_line_width=0, text=[fmt_r(v) for v in top8_ori[::-1].values],
            textposition="outside", textfont=dict(size=10.5),
            customdata=[v / total_ori_exec if total_ori_exec else 0 for v in top8_ori[::-1].values],
            hovertemplate="<b>%{y}</b><br>R$ %{x:,.0f} · %{customdata:.1%} do total<extra></extra>")
        st.plotly_chart(finish_hbar(fig_ori_exec, top8_ori.values, height=max(280, 34 * len(top8_ori))),
                         use_container_width=True, config=PLOTLY_CONFIG)
    else:
        empty_state("👥", "Coluna de Originador não encontrada — sem dados para este gráfico.")
    st.markdown("</div>", unsafe_allow_html=True)

    # ---- Slide 3 — Mix de Seguradoras e Produtos ----
    st.markdown("""<div class="exec-slide">
    <div class="exec-slide-title"><span class="exec-slide-num">3</span>Mix de Seguradoras e Produtos</div>""",
                unsafe_allow_html=True)
    if combined_tx is not None:
        tem_seg = (combined_tx["Seguradora"] != "Não informado").any()
        tem_prod = (combined_tx["Produto"] != "Não informado").any()
    else:
        tem_seg = tem_prod = False

    if tem_seg or tem_prod:
        frases = []
        if tem_seg:
            rk_seg_exec = combined_tx.groupby("Seguradora")["Valor"].sum().sort_values(ascending=False)
            tot_seg_exec = rk_seg_exec.sum()
            frases.append(f"<b>{rk_seg_exec.index[0]}</b> é a maior seguradora parceira, com "
                           f"{rk_seg_exec.iloc[0] / tot_seg_exec:.0%} da produção")
        if tem_prod:
            rk_prod_exec = combined_tx.groupby("Produto")["Valor"].sum().sort_values(ascending=False)
            tot_prod_exec = rk_prod_exec.sum()
            frases.append(f"<b>{rk_prod_exec.index[0]}</b> é a linha de produto mais forte, com "
                           f"{rk_prod_exec.iloc[0] / tot_prod_exec:.0%} do total")
        st.markdown(f'<div class="exec-headline">{"; ".join(frases)}.</div>', unsafe_allow_html=True)

        m1, m2 = st.columns(2)
        if tem_seg:
            top6_seg = rk_seg_exec.head(6)
            fig_seg_exec = px.bar(top6_seg[::-1], orientation="h", labels={"value": "R$", "Seguradora": ""},
                                   color=top6_seg[::-1].values, color_continuous_scale=[CH_BLUE_LIGHT, CH_BLUE])
            fig_seg_exec.update_layout(showlegend=False, coloraxis_showscale=False, height=280,
                                       title="Top 6 Seguradoras")
            fig_seg_exec.update_traces(marker_line_width=0, text=[fmt_r(v) for v in top6_seg[::-1].values],
                                        textposition="outside", textfont=dict(size=10),
                                        hovertemplate="<b>%{y}</b><br>R$ %{x:,.0f}<extra></extra>")
            m1.plotly_chart(finish_hbar(fig_seg_exec, top6_seg.values, height=280),
                             use_container_width=True, config=PLOTLY_CONFIG)
        else:
            m1.info("Sem coluna de Seguradora identificada nesta planilha.")
        if tem_prod:
            top6_prod = rk_prod_exec.head(6)
            fig_prod_exec = px.bar(top6_prod[::-1], orientation="h", labels={"value": "R$", "Produto": ""},
                                    color=top6_prod[::-1].values, color_continuous_scale=[CH_PURPLE, CH_PINK])
            fig_prod_exec.update_layout(showlegend=False, coloraxis_showscale=False, height=280,
                                        title="Top 6 Produtos")
            fig_prod_exec.update_traces(marker_line_width=0, text=[fmt_r(v) for v in top6_prod[::-1].values],
                                         textposition="outside", textfont=dict(size=10),
                                         hovertemplate="<b>%{y}</b><br>R$ %{x:,.0f}<extra></extra>")
            m2.plotly_chart(finish_hbar(fig_prod_exec, top6_prod.values, height=280),
                             use_container_width=True, config=PLOTLY_CONFIG)
        else:
            m2.info("Sem coluna de Produto identificada nesta planilha.")
    else:
        empty_state("🧩", "Colunas de Seguradora/Produto não encontradas — sem dados para este gráfico.")
    st.markdown("</div>", unsafe_allow_html=True)

    # ---- Slide 4 — Rentabilidade (ponte Receita → Resultado) ----
    st.markdown("""<div class="exec-slide">
    <div class="exec-slide-title"><span class="exec-slide-num">4</span>Rentabilidade da Produção</div>""",
                unsafe_allow_html=True)
    if margem_trend_delta is not None and margem_trend_delta > 0.02:
        trend_txt = f"em melhora (+{margem_trend_delta:.1%} p.p. entre a 1ª e a 2ª metade do período)"
    elif margem_trend_delta is not None and margem_trend_delta < -0.02:
        trend_txt = f"em queda ({margem_trend_delta:+.1%} p.p. entre a 1ª e a 2ª metade do período)"
    else:
        trend_txt = "estável ao longo do período"
    st.markdown(f"""<div class="exec-headline">De cada R$ 1,00 faturado, <b>{margem_lucro:.1%}</b> vira resultado
    líquido — margem {trend_txt}. Custos diretos consomem {custos_ratio:.0%} da receita e despesas administrativas
    mais {despesas_ratio:.0%}.</div>""", unsafe_allow_html=True)

    fig_water = go.Figure(go.Waterfall(
        orientation="v",
        measure=["absolute", "relative", "total", "relative", "total"],
        x=["Receita Bruta", "Custos Diretos", "Margem de Contribuição", "Despesas", "Resultado Operacional"],
        y=[receita_total, -custos_totais, 0, -despesas_totais, 0],
        text=[fmt_r(receita_total), fmt_r(-custos_totais), fmt_r(margem_contribuicao),
              fmt_r(-despesas_totais), fmt_r(resultado_operacional)],
        textposition="outside", textfont=dict(size=11),
        connector=dict(line=dict(color="rgba(140,150,160,.5)", width=1.2)),
        increasing=dict(marker=dict(color=CH_BLUE)),
        decreasing=dict(marker=dict(color=CH_DANGER)),
        totals=dict(marker=dict(color=CH_SUCCESS if resultado_operacional >= 0 else CH_DANGER)),
        hovertemplate="<b>%{x}</b><br>R$ %{y:,.0f}<extra></extra>",
    ))
    st.plotly_chart(style_fig(fig_water, height=420), use_container_width=True, config=PLOTLY_CONFIG)
    st.markdown("</div>", unsafe_allow_html=True)
